#!/usr/bin/env python3
"""
Script to examine the formula in cell H86 of worksheet M-01
"""
import sys

try:
    # Try openpyxl first for .xlsb support
    from openpyxl import load_workbook

    wb = load_workbook('/home/user/fisherman/temiz40b.xlsb', data_only=False)

    # Get worksheet M-01
    if 'M-01' in wb.sheetnames:
        ws = wb['M-01']

        # Get cell H86
        cell = ws['H86']
        print(f"Cell H86 Formula: {cell.value}")
        print(f"Cell H86 Data Type: {type(cell.value)}")

        # If it has a formula, show it
        if hasattr(cell, 'formula'):
            print(f"Cell H86 Raw Formula: {cell.formula}")

        # Show some sample data from columns C and D to understand the structure
        print("\n--- Sample data from columns C and D (rows 1-20) ---")
        for row in range(1, 21):
            c_val = ws[f'C{row}'].value
            d_val = ws[f'D{row}'].value
            print(f"Row {row}: C={c_val}, D={d_val}")

        # Show data around row 86 to understand context
        print("\n--- Data around row 86 (rows 80-90) ---")
        for row in range(80, 91):
            c_val = ws[f'C{row}'].value
            d_val = ws[f'D{row}'].value
            h_val = ws[f'H{row}'].value
            print(f"Row {row}: C={c_val}, D={d_val}, H={h_val}")

        # Look for LTQ values in column C
        print("\n--- Searching for 'LTQ' in column C ---")
        ltq_rows = []
        for row in range(1, ws.max_row + 1):
            c_val = ws[f'C{row}'].value
            if c_val and 'LTQ' in str(c_val):
                d_val = ws[f'D{row}'].value
                ltq_rows.append((row, c_val, d_val))
                if len(ltq_rows) <= 10:  # Show first 10
                    print(f"Row {row}: C={c_val}, D={d_val}")

        print(f"\nTotal rows with 'LTQ' in column C: {len(ltq_rows)}")

        # Calculate what the sum should be
        ltq_sum = sum(float(d) if d and str(d).replace('.','').replace('-','').isdigit() else 0
                      for _, _, d in ltq_rows)
        print(f"Expected sum of column D where C='LTQ': {ltq_sum}")

    else:
        print(f"Worksheet 'M-01' not found. Available sheets: {wb.sheetnames}")

except ImportError as e:
    print(f"ImportError: {e}")
    print("Trying alternative method with pyxlsb for .xlsb files...")

    try:
        from pyxlsb import open_workbook

        with open_workbook('/home/user/fisherman/temiz40b.xlsb') as wb:
            # Get worksheet M-01
            sheet_names = wb.sheets
            print(f"Available sheets: {sheet_names}")

            with wb.get_sheet('M-01') as ws:
                print("\n--- Reading M-01 worksheet ---")

                # Read all rows and store data
                rows_data = []
                for row in ws.rows():
                    rows_data.append([cell.v if cell else None for cell in row])

                print(f"Total rows: {len(rows_data)}")

                # Show headers (row 1)
                if rows_data:
                    print(f"\nRow 1 (Headers): {rows_data[0][:10]}")

                # Show row 86 (index 85)
                if len(rows_data) > 85:
                    print(f"\nRow 86: C={rows_data[85][2] if len(rows_data[85]) > 2 else None}, "
                          f"D={rows_data[85][3] if len(rows_data[85]) > 3 else None}, "
                          f"H={rows_data[85][7] if len(rows_data[85]) > 7 else None}")

                # Find LTQ rows
                ltq_rows = []
                for idx, row in enumerate(rows_data):
                    if len(row) > 2 and row[2] and 'LTQ' in str(row[2]):
                        c_val = row[2]
                        d_val = row[3] if len(row) > 3 else None
                        ltq_rows.append((idx+1, c_val, d_val))
                        if len(ltq_rows) <= 10:
                            print(f"Row {idx+1}: C={c_val}, D={d_val}")

                print(f"\nTotal rows with 'LTQ' in column C: {len(ltq_rows)}")

                # Note: pyxlsb doesn't read formulas, only values
                print("\nNote: pyxlsb cannot read formulas, only values.")
                print("We'll need to use the .xlsx version or extract formulas differently.")

    except ImportError as e2:
        print(f"pyxlsb also not available: {e2}")
        print("Need to install required libraries or use .xlsx version")
    except Exception as e2:
        print(f"Error with pyxlsb: {e2}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
