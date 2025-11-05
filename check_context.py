#!/usr/bin/env python3
"""
Script to check the context cells to understand the formula issue
"""
from openpyxl import load_workbook

print("Opening temiz40.xlsx...")
wb = load_workbook('/home/user/fisherman/temiz40.xlsx', data_only=False)
ws = wb['M-01']

print("="*60)
print("KEY CELLS:")
print("="*60)
print(f"G13: {ws['G13'].value}")
print(f"G86: {ws['G86'].value}")
print(f"G87: {ws['G87'].value}")
print(f"H86: {ws['H86'].value}")

print("\n" + "="*60)
print("COLUMN G CONTEXT (rows 80-92):")
print("="*60)
for row in range(80, 93):
    print(f"G{row}: {ws[f'G{row}'].value}")

print("\n" + "="*60)
print("UNDERSTANDING THE CURRENT FORMULA:")
print("="*60)
print(f"Current H86 formula: {ws['H86'].value}")
print("\nThis formula does:")
print("1. Checks if G86 is empty - if yes, returns empty")
print("2. Otherwise, COUNTS rows where:")
print(f"   - Column C equals G13 (which is: {ws['G13'].value})")
print(f"   - Column D is >= G86 (which is: {ws['G86'].value})")
print(f"   - Column D is < G87 (which is: {ws['G87'].value})")

print("\n" + "="*60)
print("THE PROBLEM:")
print("="*60)
print("COUNTIFS counts the NUMBER of rows matching criteria.")
print("It does NOT sum the values from column D.")
print("\nIf you want to SUM all values in column D where C='LTQ',")
print("you need a SUMIF formula instead.")

print("\n" + "="*60)
print("VERIFICATION - Let's check what the formula should return:")
print("="*60)

# Find all LTQ rows
ltq_data = []
for row in range(15, 2016):
    c_val = ws[f'C{row}'].value
    d_val = ws[f'D{row}'].value
    if c_val == 'LTQ' and d_val is not None:
        # Handle formula cells in column D
        if isinstance(d_val, str) and d_val.startswith('='):
            # This is a formula, we'd need to evaluate it
            # For now, let's use data_only mode
            ltq_data.append((row, c_val, d_val, "FORMULA"))
        else:
            ltq_data.append((row, c_val, d_val, "VALUE"))

print(f"\nFound {len(ltq_data)} rows with 'LTQ' in column C:")
for row, c, d, dtype in ltq_data[:15]:
    print(f"  Row {row}: C={c}, D={d} ({dtype})")

# Now load with data_only=True to get calculated values
wb_calc = load_workbook('/home/user/fisherman/temiz40.xlsx', data_only=True)
ws_calc = wb_calc['M-01']

print("\n" + "="*60)
print("CALCULATED VALUES (with formulas evaluated):")
print("="*60)

ltq_sum = 0
ltq_count = 0
g86_val = ws_calc['G86'].value
g87_val = ws_calc['G87'].value

print(f"G86 value: {g86_val}")
print(f"G87 value: {g87_val}")

for row in range(15, 2016):
    c_val = ws_calc[f'C{row}'].value
    d_val = ws_calc[f'D{row}'].value
    if c_val == 'LTQ' and d_val is not None:
        try:
            d_num = float(d_val)
            ltq_sum += d_num
            ltq_count += 1
            # Check if this value falls in the G86-G87 bin
            if g86_val is not None and g87_val is not None:
                if d_num >= float(g86_val) and d_num < float(g87_val):
                    print(f"  Row {row}: C={c_val}, D={d_num} (IN BIN [{g86_val}, {g87_val}))")
        except (ValueError, TypeError):
            pass

print(f"\nTotal SUM of all LTQ values: {ltq_sum}")
print(f"Total COUNT of LTQ rows: {ltq_count}")
print(f"\nH86 calculated value: {ws_calc['H86'].value}")

print("\n" + "="*60)
print("RECOMMENDED FIX:")
print("="*60)
print("\nIf you want to SUM all column D values where C='LTQ':")
print("  =SUMIF($C$15:$C$2015,\"LTQ\",$D$15:$D$2015)")
print("  or")
print("  =SUMIF($C$15:$C$2015,G$13,$D$15:$D$2015)")
print("\nIf you want to keep the frequency distribution but fix it:")
print("  (The current formula is actually correct for frequency distribution)")
