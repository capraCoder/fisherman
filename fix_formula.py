#!/usr/bin/env python3
"""
Script to fix the formula in cell H86 of worksheet M-01
"""
from openpyxl import load_workbook
import shutil

# Create backup first
print("Creating backup of temiz40.xlsx...")
shutil.copy('/home/user/fisherman/temiz40.xlsx', '/home/user/fisherman/temiz40.xlsx.backup')

print("Opening temiz40.xlsx...")
wb = load_workbook('/home/user/fisherman/temiz40.xlsx')
ws = wb['M-01']

print("\n" + "="*60)
print("CURRENT STATE:")
print("="*60)
print(f"H86 current formula: {ws['H86'].value}")

# Fix the formula
old_formula = ws['H86'].value
new_formula = "=SUMIF($C$15:$C$2015,G$13,$D$15:$D$2015)"

print("\n" + "="*60)
print("APPLYING FIX:")
print("="*60)
print(f"Old formula: {old_formula}")
print(f"New formula: {new_formula}")

ws['H86'] = new_formula

print("\n" + "="*60)
print("VERIFICATION:")
print("="*60)
print(f"H86 new formula: {ws['H86'].value}")

# Save the file
print("\nSaving temiz40.xlsx...")
wb.save('/home/user/fisherman/temiz40.xlsx')

print("\n" + "="*60)
print("SUCCESS!")
print("="*60)
print("The formula in H86 has been updated.")
print(f"Expected result: 450.0 (sum of all D values where C='LTQ')")
print("\nBackup saved as: temiz40.xlsx.backup")
