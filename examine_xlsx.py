#!/usr/bin/env python3
"""
Script to examine the .xlsx file and get the formula from H86
"""
from openpyxl import load_workbook

print("Opening temiz40.xlsx...")
wb = load_workbook('/home/user/fisherman/temiz40.xlsx', data_only=False)

# Get worksheet M-01
if 'M-01' in wb.sheetnames:
    ws = wb['M-01']
    print(f"Opened worksheet: M-01\n")

    # Get cell H86
    cell_h86 = ws['H86']
    print("="*60)
    print("CELL H86 INFORMATION:")
    print("="*60)
    print(f"Value: {cell_h86.value}")
    print(f"Data Type: {type(cell_h86.value)}")

    # Check if it's a formula
    if cell_h86.value and isinstance(cell_h86.value, str) and cell_h86.value.startswith('='):
        print(f"Formula: {cell_h86.value}")
    else:
        print("No formula detected (might be a plain value or empty)")

    # Check the internal formula attribute
    if hasattr(cell_h86, '_value'):
        print(f"Internal _value: {cell_h86._value}")

    print("\n" + "="*60)
    print("CELL H1 (FOR COMPARISON):")
    print("="*60)
    cell_h1 = ws['H1']
    print(f"G1 Value: {ws['G1'].value}")
    print(f"H1 Value: {cell_h1.value}")
    if cell_h1.value and isinstance(cell_h1.value, str) and cell_h1.value.startswith('='):
        print(f"H1 Formula: {cell_h1.value}")

    # Show some context around row 86
    print("\n" + "="*60)
    print("CONTEXT - Rows 80-92, Column H:")
    print("="*60)
    for row in range(80, 93):
        cell = ws[f'H{row}']
        value = cell.value
        formula = ""
        if value and isinstance(value, str) and value.startswith('='):
            formula = f" [Formula: {value}]"
        print(f"H{row}: {value}{formula}")

    # Check column structure
    print("\n" + "="*60)
    print("SAMPLE DATA STRUCTURE (rows with LTQ):")
    print("="*60)
    ltq_rows = [15, 23, 31, 39, 47, 55, 63, 71]
    for row in ltq_rows:
        c_val = ws[f'C{row}'].value
        d_val = ws[f'D{row}'].value
        print(f"Row {row}: C={c_val}, D={d_val}")

    # Look for other formulas in column H that might give us a clue
    print("\n" + "="*60)
    print("OTHER FORMULAS IN COLUMN H (first 10 found):")
    print("="*60)
    formula_count = 0
    for row in range(1, min(100, ws.max_row + 1)):
        cell = ws[f'H{row}']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"H{row}: {cell.value}")
            formula_count += 1
            if formula_count >= 10:
                break

else:
    print(f"Worksheet 'M-01' not found. Available sheets: {wb.sheetnames}")
