#!/usr/bin/env python3
"""
Script to verify the fix works correctly
"""
from openpyxl import load_workbook

print("Verifying the fix...")
print("\n" + "="*60)
print("LOADING WITH data_only=True TO SEE CALCULATED VALUES:")
print("="*60)

wb = load_workbook('/home/user/fisherman/temiz40.xlsx', data_only=True)
ws = wb['M-01']

print(f"\nG13 value (FAO code): {ws['G13'].value}")
print(f"H86 calculated value: {ws['H86'].value}")

print("\n" + "="*60)
print("LOADING WITH data_only=False TO SEE FORMULAS:")
print("="*60)

wb2 = load_workbook('/home/user/fisherman/temiz40.xlsx', data_only=False)
ws2 = wb2['M-01']

print(f"\nH86 formula: {ws2['H86'].value}")

print("\n" + "="*60)
print("MANUAL CALCULATION VERIFICATION:")
print("="*60)

# Manually calculate the expected sum
ltq_sum = 0
ltq_rows = []
for row in range(15, 2016):
    c_val = ws[f'C{row}'].value
    d_val = ws[f'D{row}'].value
    if c_val == 'LTQ' and d_val is not None:
        try:
            d_num = float(d_val)
            ltq_sum += d_num
            ltq_rows.append((row, d_num))
        except (ValueError, TypeError):
            pass

print(f"\nRows with 'LTQ' in column C:")
for row, d_val in ltq_rows:
    print(f"  Row {row}: D={d_val}")

print(f"\nManually calculated sum: {ltq_sum}")
print(f"Expected: 450.0")
print(f"Match: {ltq_sum == 450.0}")

print("\n" + "="*60)
print("NOTE ABOUT .xlsb FILE:")
print("="*60)
print("The .xlsb file cannot be directly edited with Python libraries.")
print("To update temiz40b.xlsb:")
print("1. Open temiz40.xlsx in Microsoft Excel")
print("2. Navigate to worksheet M-01, cell H86")
print("3. Verify the formula is: =SUMIF($C$15:$C$2015,G$13,$D$15:$D$2015)")
print("4. Save As > Excel Binary Workbook (.xlsb)")
print("5. Replace the existing temiz40b.xlsb file")
